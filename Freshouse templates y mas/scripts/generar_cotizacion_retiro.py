#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generador de cotizacion en Excel — Conjunto Residencial El Retiro
Mantenimiento de Fachada | COT-RET-2026-001
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "/Users/andresmontoyaosorio/Documents/TRABAJO/CLAUDE_SEX/CLAUDE/cotizacion_conjunto_el_retiro.xlsx"

# ---------------------------------------------------------------------------
# Definicion de datos
# ---------------------------------------------------------------------------

GRUPOS = [
    {
        "codigo": "GRP-1",
        "nombre": "PRELIMINARES",
        "items": [
            ("1.1", "Campamento de obra",                                          "UND", 1,       1_500_000),
            ("1.2", "Señalización y seguridad",                                    "GLB", 1,       2_200_000),
        ],
    },
    {
        "codigo": "GRP-2",
        "nombre": "LIMPIEZA Y ALISTAMIENTO",
        "items": [
            ("2.1", "Hidrolavado hipoclorito 13%, cepillado, retiro baja adherencia", "M2", 2852.21, 8_500),
            ("2.2", "Reposición graniplast (imprimante + mortero + textura)",         "M2",   85.57, 95_000),
        ],
    },
    {
        "codigo": "GRP-3",
        "nombre": "PINTURA E IMPERMEABILIZACIÓN",
        "items": [
            ("3.1", "Sellos ventanería perimetral — Soudal Multibond",              "ML",   693.20, 28_000),
            ("3.2", "Koraza Pro 550 a 2 manos",                                     "M2",  2505.28, 22_000),
            ("3.3", "Dilataciones Pintucofill 7 años",                              "ML",  1690.88, 14_500),
            ("3.4", "Alfajías — masillado + Koraza Pro 550",                        "ML",   250.08, 18_000),
            ("3.5", "Vinilo tipo 1 parqueadero",                                    "M2",   313.88, 16_500),
            ("3.6", "Esmalte Pintulux 3 en 1 carpintería metálica",                "M2",   260.00, 32_000),
        ],
    },
    {
        "codigo": "GRP-4",
        "nombre": "TERMINADO Y LIMPIEZA",
        "items": [
            ("4.1", "Aseo general y retiro de pintura",                             "UND", 1,       1_200_000),
        ],
    },
]

SUBTOTAL_SIN_IVA = 154_316_915
IVA_PCT          = 0.19
IVA_VALOR        = 29_320_214
TOTAL_CON_IVA    = 183_637_129

# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------

def color_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def thin_border():
    thin = Side(style="thin", color="BDBDBD")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def header_border():
    medium = Side(style="medium", color="757575")
    return Border(left=medium, right=medium, top=medium, bottom=medium)

FILL_ENCABEZADO   = color_fill("1B5E20")   # verde oscuro — encabezado principal
FILL_GRUPO        = color_fill("EEEEEE")   # gris claro — filas de grupo
FILL_SUBTOTAL     = color_fill("E8F5E9")   # verde muy claro — subtotales
FILL_IVA_ROW      = color_fill("F5F5F5")   # gris muy claro — IVA
FILL_TOTAL        = color_fill("2E7D32")   # verde oscuro — total general
FILL_COL_HEADER   = color_fill("424242")   # gris oscuro — cabecera de columnas

FONT_BLANCO_BOLD  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_BOLD         = Font(name="Calibri", bold=True, size=10)
FONT_NORMAL       = Font(name="Calibri", size=10)
FONT_TITULO       = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
FONT_SUBTITULO    = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

ALIGN_CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_RIGHT       = Alignment(horizontal="right",  vertical="center")

FMT_MONEDA        = '"$"#,##0'        # moneda colombiana sin decimales
FMT_CANT          = '#,##0.00'        # cantidad 2 decimales
FMT_ENTERO        = '#,##0'

# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------

def set_cell(ws, row, col, value, font=None, fill=None, alignment=None,
             border=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:            cell.font            = font
    if fill:            cell.fill            = fill
    if alignment:       cell.alignment       = alignment
    if border:          cell.border          = border
    if number_format:   cell.number_format   = number_format
    return cell

def merge_and_set(ws, row, col_start, col_end, value, font=None, fill=None,
                  alignment=None, border=None):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=value)
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border
    # rellenar celdas mergeadas con mismo fill para bordes limpios
    for c in range(col_start + 1, col_end + 1):
        aux = ws.cell(row=row, column=c)
        if fill: aux.fill = fill
    return cell

# ---------------------------------------------------------------------------
# Construccion del workbook
# ---------------------------------------------------------------------------

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Cotizacion"

# Columnas: A=Item, B=Descripcion, C=Und, D=Cantidad, E=V.Unitario, F=V.Total
COL_WIDTHS = {1: 8, 2: 55, 3: 8, 4: 12, 5: 18, 6: 18}
for col, width in COL_WIDTHS.items():
    ws.column_dimensions[get_column_letter(col)].width = width

ws.row_dimensions[1].height = 40
ws.row_dimensions[2].height = 18
ws.row_dimensions[3].height = 18
ws.row_dimensions[4].height = 18
ws.row_dimensions[5].height = 22

# --- Fila 1: Titulo principal ---
merge_and_set(ws, 1, 1, 6,
    "CONJUNTO RESIDENCIAL EL RETIRO\nMANTENIMIENTO DE FACHADA",
    font=FONT_TITULO, fill=FILL_ENCABEZADO, alignment=ALIGN_CENTER)

# --- Fila 2: Datos del documento ---
merge_and_set(ws, 2, 1, 3, "Cotizacion No.:",
    font=FONT_SUBTITULO, fill=FILL_ENCABEZADO, alignment=ALIGN_LEFT)
merge_and_set(ws, 2, 4, 6, "COT-RET-2026-001",
    font=FONT_SUBTITULO, fill=FILL_ENCABEZADO, alignment=ALIGN_LEFT)

# --- Fila 3: Fecha ---
merge_and_set(ws, 3, 1, 3, "Fecha:",
    font=FONT_SUBTITULO, fill=FILL_ENCABEZADO, alignment=ALIGN_LEFT)
merge_and_set(ws, 3, 4, 6, "2026-04-19",
    font=FONT_SUBTITULO, fill=FILL_ENCABEZADO, alignment=ALIGN_LEFT)

# --- Fila 4: Empresa ---
merge_and_set(ws, 4, 1, 3, "Elaborado por:",
    font=FONT_SUBTITULO, fill=FILL_ENCABEZADO, alignment=ALIGN_LEFT)
merge_and_set(ws, 4, 4, 6, "FreshHouse S.A.S — Mantenimiento y Restauracion de Fachadas",
    font=FONT_SUBTITULO, fill=FILL_ENCABEZADO, alignment=ALIGN_LEFT)

# --- Fila 5: Encabezados de columna ---
HEADERS = ["ITEM", "DESCRIPCION", "UND", "CANTIDAD", "V. UNITARIO", "V. TOTAL"]
for col, header in enumerate(HEADERS, start=1):
    set_cell(ws, 5, col, header,
             font=FONT_BLANCO_BOLD, fill=FILL_COL_HEADER,
             alignment=ALIGN_CENTER, border=header_border())

# ---------------------------------------------------------------------------
# Filas de datos
# ---------------------------------------------------------------------------

current_row = 6

for grupo in GRUPOS:
    # --- Fila de grupo ---
    ws.row_dimensions[current_row].height = 20
    merge_and_set(ws, current_row, 1, 6,
        f"{grupo['codigo']} — {grupo['nombre']}",
        font=FONT_BOLD, fill=FILL_GRUPO, alignment=ALIGN_LEFT)
    # borde para toda la fila mergeada
    for c in range(1, 7):
        ws.cell(row=current_row, column=c).border = thin_border()
    current_row += 1

    subtotal_grupo = 0

    for item in grupo["items"]:
        codigo, desc, und, cantidad, v_unitario = item
        v_total = round(cantidad * v_unitario)
        subtotal_grupo += v_total
        ws.row_dimensions[current_row].height = 18

        set_cell(ws, current_row, 1, codigo,
                 font=FONT_NORMAL, alignment=ALIGN_CENTER, border=thin_border())
        set_cell(ws, current_row, 2, desc,
                 font=FONT_NORMAL, alignment=ALIGN_LEFT, border=thin_border())
        set_cell(ws, current_row, 3, und,
                 font=FONT_NORMAL, alignment=ALIGN_CENTER, border=thin_border())
        set_cell(ws, current_row, 4, cantidad,
                 font=FONT_NORMAL, alignment=ALIGN_RIGHT, border=thin_border(),
                 number_format=FMT_CANT)
        set_cell(ws, current_row, 5, v_unitario,
                 font=FONT_NORMAL, alignment=ALIGN_RIGHT, border=thin_border(),
                 number_format=FMT_MONEDA)
        set_cell(ws, current_row, 6, v_total,
                 font=FONT_NORMAL, alignment=ALIGN_RIGHT, border=thin_border(),
                 number_format=FMT_MONEDA)
        current_row += 1

    # --- Fila de subtotal del grupo ---
    ws.row_dimensions[current_row].height = 18
    merge_and_set(ws, current_row, 1, 5,
        f"Subtotal {grupo['codigo']} — {grupo['nombre']}",
        font=FONT_BOLD, fill=FILL_SUBTOTAL, alignment=ALIGN_RIGHT)
    for c in range(1, 6):
        ws.cell(row=current_row, column=c).border = thin_border()
    set_cell(ws, current_row, 6, subtotal_grupo,
             font=FONT_BOLD, fill=FILL_SUBTOTAL,
             alignment=ALIGN_RIGHT, border=thin_border(),
             number_format=FMT_MONEDA)
    current_row += 1

# --- Fila separadora ---
current_row += 1

# --- Subtotal sin IVA ---
ws.row_dimensions[current_row].height = 18
merge_and_set(ws, current_row, 1, 5,
    "SUBTOTAL SIN IVA",
    font=FONT_BOLD, fill=FILL_IVA_ROW, alignment=ALIGN_RIGHT)
for c in range(1, 6):
    ws.cell(row=current_row, column=c).border = thin_border()
set_cell(ws, current_row, 6, SUBTOTAL_SIN_IVA,
         font=FONT_BOLD, fill=FILL_IVA_ROW,
         alignment=ALIGN_RIGHT, border=thin_border(),
         number_format=FMT_MONEDA)
current_row += 1

# --- IVA 19% ---
ws.row_dimensions[current_row].height = 18
merge_and_set(ws, current_row, 1, 5,
    "IVA 19%",
    font=FONT_BOLD, fill=FILL_IVA_ROW, alignment=ALIGN_RIGHT)
for c in range(1, 6):
    ws.cell(row=current_row, column=c).border = thin_border()
set_cell(ws, current_row, 6, IVA_VALOR,
         font=FONT_BOLD, fill=FILL_IVA_ROW,
         alignment=ALIGN_RIGHT, border=thin_border(),
         number_format=FMT_MONEDA)
current_row += 1

# --- TOTAL GENERAL ---
ws.row_dimensions[current_row].height = 24
merge_and_set(ws, current_row, 1, 5,
    "TOTAL GENERAL CON IVA",
    font=FONT_BLANCO_BOLD, fill=FILL_TOTAL, alignment=ALIGN_RIGHT)
for c in range(1, 6):
    ws.cell(row=current_row, column=c).border = header_border()
set_cell(ws, current_row, 6, TOTAL_CON_IVA,
         font=FONT_BLANCO_BOLD, fill=FILL_TOTAL,
         alignment=ALIGN_RIGHT, border=header_border(),
         number_format=FMT_MONEDA)
current_row += 2

# --- Notas al pie ---
ws.row_dimensions[current_row].height = 14
merge_and_set(ws, current_row, 1, 6,
    "NOTAS:",
    font=Font(name="Calibri", bold=True, size=9, color="424242"),
    alignment=ALIGN_LEFT)
current_row += 1

notas = [
    "1. Precios incluyen mano de obra, materiales, herramienta menor y APU conforme a mercado Eje Cafetero 2026.",
    "2. Validez de la oferta: 30 dias calendario.",
    "3. Forma de pago: anticipo 40% — 40% avance 50% — 20% entrega final.",
    "4. Los trabajos en altura se realizan con certificacion SGSST vigente y dotacion completa.",
    "5. Cualquier actividad adicional no contemplada en el presente alcance sera cotizada por separado.",
]
for nota in notas:
    ws.row_dimensions[current_row].height = 13
    merge_and_set(ws, current_row, 1, 6, nota,
        font=Font(name="Calibri", size=9, color="616161"),
        alignment=ALIGN_LEFT)
    current_row += 1

# ---------------------------------------------------------------------------
# Inmovilizar paneles en fila 6 (bajo encabezados)
# ---------------------------------------------------------------------------
ws.freeze_panes = "A6"

# ---------------------------------------------------------------------------
# Guardar
# ---------------------------------------------------------------------------
wb.save(OUTPUT_PATH)
print(f"Archivo generado: {OUTPUT_PATH}")
