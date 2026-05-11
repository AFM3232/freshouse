#!/usr/bin/env python3
"""Genera XLSX editable para cotización Torre Providencia II (2 alternativas)."""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Paleta ────────────────────────────────────────────────────────────────────
C_BLUE_DARK  = "0F2D5C"   # encabezado principal
C_BLUE_MID   = "1A52A8"   # grupo header
C_BLUE_LIGHT = "EEF3FD"   # fila alternada
C_ORANGE     = "F97316"   # totales grand
C_ORANGE_BG  = "FFF3E8"   # IVA row bg
C_WHITE      = "FFFFFF"
C_TOTAL_BG   = "E4EDF8"   # subtotal bg
C_GRAND_BG   = "2563EB"   # grand total bg
C_GRAND_FG   = "FFFFFF"
C_GRP_BG     = "1A52A8"
C_GRP_FG     = "FFFFFF"
C_LOCKED_BG  = "F0F4FF"   # celdas calculadas

FMT_COP  = '"$"#,##0.00'
FMT_PCT  = '0"%"'
FMT_NUM  = '#,##0.00'

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

def border_thin():
    s = Side(style="thin", color="C5D0E6")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium_top():
    thin = Side(style="thin",   color="C5D0E6")
    med  = Side(style="medium", color="1A52A8")
    return Border(left=thin, right=thin, top=med, bottom=thin)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def right():
    return Alignment(horizontal="right", vertical="center")

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ── Datos ─────────────────────────────────────────────────────────────────────
ITEMS = [
    # (item, descripcion, und, cantidad, v_unitario, iva_pct, grupo)
    # GRP-1
    ("GRP-1", "Preliminares", None, None, None, None, True),
    ("1.1", "Campamento de obra", "UND", 1.00, 1200000.00, 19, False),
    ("1.2", "Señalización y seguridad de obra", "GLB", 1.00, 400000.00, 19, False),
    # GRP-2
    ("GRP-2", "Limpieza y Alistamiento de Superficies", None, None, None, None, True),
    ("2.1", "Hidrolavado a presión con solución de hipoclorito al 13%. Cepillado y enjuague con hidrolavadora 1880psi. Retiro de material con ampollamientos o baja adherencia", "M2", 2852.21, 4522.00, 19, False),
    ("2.2", "Reposición de graniplast: imprimante, mortero de reparación impermeabilizado, segunda imprimación y textura graniplast", "M2", 85.57, 52360.00, 19, False),
    # GRP-3
    ("GRP-3", "Obras de Pintura en Muros de Fachada e Impermeabilización", None, None, None, None, True),
    ("3.1", "Revisión y recambio integral de sellos de ventanería perimetral. Suministro y aplicación Soudal Multibond", "ML", 693.20, 8687.00, 19, False),
    ("3.2", "Suministro y aplicación de pintura Koraza Pro 550 a 2 manos según ficha técnica", "M2", 2505.28, 23205.00, 19, False),
    ("3.3", "Tratamiento de dilataciones con Pintucofill 7 años de Pintuco", "ML", 1690.88, 8687.00, 19, False),
    ("3.4", "Tratamiento en alfajías y aplicación de Koraza Pro 550", "ML", 250.08, 10115.00, 19, False),
    ("3.5", "Suministro y aplicación de vinilo tipo 1 en área de parqueadero", "M2", 313.88, 17850.00, 19, False),
    ("3.6", "Suministro y aplicación de esmalte Pintulux 3 en 1 sobre carpintería metálica", "M2", 260.00, 27370.00, 19, False),
    # GRP-4
    ("GRP-4", "Terminado y Limpieza", None, None, None, None, True),
    ("4.1", "Aseo general y retiro de pintura", "UND", 1.00, 3094000.00, 19, False),
]


def build_sheet(ws, alt_num):
    # ── Título ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = f"COTIZACIÓN TORRE PROVIDENCIA II — ALTERNATIVA {alt_num}"
    ws["A1"].font      = font(bold=True, color=C_GRAND_FG, size=13)
    ws["A1"].fill      = fill(C_BLUE_DARK)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    ws["A2"] = "FreshHouse SAS · Armenia, Quindío · Cód. FH-F-02 · Válido hasta 13/04/2026"
    ws["A2"].font      = font(italic=True, color="849AB8", size=9)
    ws["A2"].fill      = fill("F8FAFF")
    ws["A2"].alignment = center()

    # ── Encabezados de columna ────────────────────────────────────────────────
    headers = ["Ítem", "Descripción", "Und", "Cantidad\n(editar)", "V. Unitario\n(editar)", "IVA %\n(editar)", "V. Total\n(calculado)"]
    cols    = ["A",    "B",           "C",   "D",                  "E",                     "F",               "G"]
    widths  = [8,       52,            8,     14,                   16,                      10,                18]

    for col_letter, header, width in zip(cols, headers, widths):
        cell = ws[f"{col_letter}3"]
        cell.value     = header
        cell.font      = font(bold=True, color=C_GRAND_FG, size=10)
        cell.fill      = fill(C_BLUE_DARK)
        cell.alignment = center()
        cell.border    = border_thin()
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[3].height = 30

    # ── IVA validation ────────────────────────────────────────────────────────
    dv = DataValidation(type="list", formula1='"0,5,19"', allow_blank=False)
    dv.sqref = "F4:F200"
    ws.add_data_validation(dv)

    # ── Filas de ítems ────────────────────────────────────────────────────────
    row = 4
    item_rows = []   # (row, is_group)

    for item_data in ITEMS:
        cod, desc, und, cant, vunit, iva_pct, is_group = item_data

        if is_group:
            ws.merge_cells(f"A{row}:F{row}")
            ws[f"A{row}"] = f"  {cod}  ·  {desc}"
            ws[f"A{row}"].font      = font(bold=True, color=C_GRP_FG, size=10)
            ws[f"A{row}"].fill      = fill(C_GRP_BG)
            ws[f"A{row}"].alignment = left()
            ws[f"A{row}"].border    = border_medium_top()

            ws[f"G{row}"]           = f'=SUMIF(A{row+1}:A{row+20},"<>",G{row+1}:G{row+20})'
            ws[f"G{row}"].font      = font(bold=True, color="F97316", size=10)
            ws[f"G{row}"].fill      = fill(C_GRP_BG)
            ws[f"G{row}"].alignment = right()
            ws[f"G{row}"].border    = border_medium_top()
            ws[f"G{row}"].number_format = FMT_COP
            ws.row_dimensions[row].height = 20
            item_rows.append((row, True))
            row += 1
        else:
            row_bg = C_WHITE if (row % 2 == 0) else C_BLUE_LIGHT

            ws[f"A{row}"] = cod
            ws[f"A{row}"].font      = font(bold=True, size=9)
            ws[f"A{row}"].fill      = fill(row_bg)
            ws[f"A{row}"].alignment = center()
            ws[f"A{row}"].border    = border_thin()

            ws[f"B{row}"] = desc
            ws[f"B{row}"].font      = font(size=9)
            ws[f"B{row}"].fill      = fill(row_bg)
            ws[f"B{row}"].alignment = left()
            ws[f"B{row}"].border    = border_thin()

            ws[f"C{row}"] = und
            ws[f"C{row}"].font      = font(size=9)
            ws[f"C{row}"].fill      = fill(row_bg)
            ws[f"C{row}"].alignment = center()
            ws[f"C{row}"].border    = border_thin()

            # Editable: cantidad
            ws[f"D{row}"] = cant
            ws[f"D{row}"].font           = font(bold=True, size=10, color="0F2D5C")
            ws[f"D{row}"].fill           = fill("FFFCE8")
            ws[f"D{row}"].alignment      = right()
            ws[f"D{row}"].border         = border_thin()
            ws[f"D{row}"].number_format  = FMT_NUM

            # Editable: v. unitario
            ws[f"E{row}"] = vunit
            ws[f"E{row}"].font           = font(bold=True, size=10, color="0F2D5C")
            ws[f"E{row}"].fill           = fill("FFFCE8")
            ws[f"E{row}"].alignment      = right()
            ws[f"E{row}"].border         = border_thin()
            ws[f"E{row}"].number_format  = FMT_COP

            # Editable: IVA %
            ws[f"F{row}"] = iva_pct
            ws[f"F{row}"].font           = font(bold=True, size=10, color="0F2D5C")
            ws[f"F{row}"].fill           = fill("FFFCE8")
            ws[f"F{row}"].alignment      = center()
            ws[f"F{row}"].border         = border_thin()
            ws[f"F{row}"].number_format  = FMT_PCT

            # Calculado: V. Total sin IVA (base)
            ws[f"G{row}"] = f"=D{row}*E{row}"
            ws[f"G{row}"].font           = font(size=10, color="0F2D5C")
            ws[f"G{row}"].fill           = fill(C_LOCKED_BG)
            ws[f"G{row}"].alignment      = right()
            ws[f"G{row}"].border         = border_thin()
            ws[f"G{row}"].number_format  = FMT_COP

            ws.row_dimensions[row].height = 40 if len(desc) > 80 else 22
            item_rows.append((row, False))
            row += 1

    # ── Filas de totales ──────────────────────────────────────────────────────
    row += 1  # separador
    subtotal_row = row

    # Recopila todas las filas de ítems (no grupo) para la fórmula
    item_data_rows = [r for r, g in item_rows if not g]
    g_refs  = "+".join([f"G{r}" for r in item_data_rows])
    iva_ref = "+".join([f"D{r}*E{r}*F{r}/100" for r in item_data_rows])

    def total_row(ws, row, label, formula, bg, fg, size=10, bold=True, border_top=False):
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"]           = label
        ws[f"A{row}"].font      = font(bold=bold, color=fg, size=size)
        ws[f"A{row}"].fill      = fill(bg)
        ws[f"A{row}"].alignment = right()
        b = border_medium_top() if border_top else border_thin()
        ws[f"A{row}"].border    = b

        ws[f"G{row}"]              = f"={formula}"
        ws[f"G{row}"].font         = font(bold=bold, color=fg, size=size)
        ws[f"G{row}"].fill         = fill(bg)
        ws[f"G{row}"].alignment    = right()
        ws[f"G{row}"].border       = b
        ws[f"G{row}"].number_format = FMT_COP
        ws.row_dimensions[row].height = 22

    total_row(ws, subtotal_row,   "COSTO DIRECTO (sin IVA)",     g_refs,                      C_TOTAL_BG, C_BLUE_DARK, border_top=True)
    total_row(ws, subtotal_row+1, "IVA (según % por ítem)",      iva_ref,                     C_ORANGE_BG, C_ORANGE)
    total_row(ws, subtotal_row+2, "VALOR TOTAL",                 f"G{subtotal_row}+G{subtotal_row+1}", C_GRAND_BG, C_GRAND_FG, size=12)

    # ── Nota instructiva ──────────────────────────────────────────────────────
    note_row = subtotal_row + 4
    ws.merge_cells(f"A{note_row}:G{note_row}")
    ws[f"A{note_row}"] = "⚠  Celdas en amarillo son editables: Cantidad · V. Unitario · IVA %.  Columna G (azul claro) es calculada — no editar."
    ws[f"A{note_row}"].font      = font(italic=True, size=9, color="849AB8")
    ws[f"A{note_row}"].alignment = left()

    # ── Freeze panes ─────────────────────────────────────────────────────────
    ws.freeze_panes = "A4"


# ── Main ──────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

for alt in [1, 2]:
    ws = wb.create_sheet(title=f"Alternativa {alt}")
    build_sheet(ws, alt)

out_path = "/Users/andresmontoyaosorio/Documents/TRABAJO/CLAUDE_SEX/CLAUDE/cotizacion_torre_providencia_II.xlsx"
wb.save(out_path)
print(f"Guardado: {out_path}")
